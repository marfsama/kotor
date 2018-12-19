#!/usr/bin/env python3

import argparse
import os
import io
import itertools

from openpyxl import Workbook, load_workbook
from tools import *
from collections import OrderedDict

class Header:
    def __init__(self, file):
        self.marker = file.read(4).decode("utf-8")
        self.version = file.read(4).decode("utf-8")

    def __str__(self):
        return """{name}: {{magic: "{marker}{version}"}}""".format(name=type(self).__name__, **vars(self))

class TabularDataFile:
    def __init__(self, column_names, rows):
        self.column_names = column_names
        self.rows = rows

    def __str__(self):
        return """{name}: {{columns: {column_names}, rows: {rows}}}""".format(name=type(self).__name__, **vars(self))

def execute_action(parsed, keyFile):
    switcher= {
        "list" : list_entries,
        "update" : update_entry,
        "extract" : extract_entry,
        "delete" : delete_entry
    }
    func = switcher.get(parsed.action, lambda: print("You need to specify one of -l, -u, -x, -d"))
    func(parsed, keyFile)

def read_byte_by_byte(file):
    """
        returns the file as a byte by byte iterator
    """
    while True:
        yield file.read(1)

def read_tokens(file):
    """
        returns tokens separated by \x00 or tab
    """
    current_column = ""
    for byte in read_byte_by_byte(file):
        if byte in [b"\x00", b"\t"]:
            yield current_column
            current_column = ""
        else:
            current_column = current_column + byte.decode("utf-8")

def read_token_list(file):
    """
        reads tokens until an empty token is encountered
    """
    for token in read_tokens(file):
        if not token:
            break
        yield token


def read_file(file_name):
    with open(file_name, "rb") as file:
        header = Header(file)
        if header.version != "V2.b":
            print ("wrong version. supported is V2.b, version is", header.version)
            return
        # skip new line
        file.read(1)
        # read column names. column names are separated by tab (0x9), a null (0x0) column name signifies the end of the names list
        column_names = list(read_token_list(file))
        # read row count and row indices
        num_rows = read32(file)
        row_indices = list(itertools.islice(read_tokens(file), num_rows))

        # read cell offsets
        cell_offsets = readlist(read16, file, num_rows * len(column_names))

        # read string data into inmemory buffer
        data_size = read16(file)
        data = file.read(data_size)
        data_stream = io.BytesIO(data)

        # parse cell contents from string data
        cell_offset_index = 0
        rows = OrderedDict() # note: use OrderedDict to keep entries and iterations in insertion order
        for row_index in row_indices:
            row = []
            for column_name in column_names:
                data_stream.seek(cell_offsets[cell_offset_index])
                token = next(read_tokens(data_stream))
                row.append(token)
                cell_offset_index = cell_offset_index+1
            rows[row_index] = row

        return TabularDataFile(column_names, rows)

def get_output_filename(parsed):
    if parsed.output:
        return parsed.output
    filename = os.path.split(parsed.input)[1]
    basename = os.path.splitext(filename)[0]
    format_extensions = {
        "excel" : "xlsx",
        "csv" : "csv"
    }
    return basename + '.' + format_extensions[parsed.format]

def extract_csv(parsed):
    tabular_data_file = read_file(parsed.input)
    output_filename = get_output_filename(parsed)
    csv_delimiter = parsed.csvsep

    print('extract csv to', output_filename)
    with open(output_filename, 'w') as output:
        output.write('index'+csv_delimiter)
        output.write(csv_delimiter.join(tabular_data_file.column_names))
        output.write('\n')
        for row, values in tabular_data_file.rows.items():
            output.write(row+ csv_delimiter)
            output.write(csv_delimiter.join(values))
            output.write('\n')


def extract_excel(parsed):
    tabular_data_file = read_file(parsed.input)
    output_filename = get_output_filename(parsed)
    print('extract excel to', output_filename)
    wb = Workbook()
    # grab the active worksheet
    ws = wb.active
    # Rows can also be appended
    ws.append(['index']+ tabular_data_file.column_names)

    for row, values in tabular_data_file.rows.items():
        ws.append([row]+values)

    # Save the file
    wb.save(output_filename)
    pass

def read_excel_file(filename):
    wb = load_workbook(filename=filename, read_only=True)
    ws = wb.active

    sheet_rows = list(ws.rows)

    # first line is column names (skip first column, which is the row index)
    column_names = [cell.value for cell in sheet_rows[0][1:]]


    # read rows, saving the row index separately
    # note: use OrderedDict to keep entries and iterations in insertion order
    rows = OrderedDict()
    for row in sheet_rows[1:]:
        row_index = row[0].value
        rows[row_index] = [cell.value for cell in row[1:]]

    return TabularDataFile(column_names, rows)

def write_2da_file(output_filename, tabular_data_file):
    print('writing 2da file:', output_filename)
    with open(output_filename, 'wb') as file:
        # write header
        file.write('2DA V2.b'.encode('utf-8'))
        # write newline
        file.write(bytes([0xa]))
        # write column names, each terminated by tab (0x9)
        for column_name in tabular_data_file.column_names:
            file.write(column_name.encode('utf-8'))
            file.write(bytes([0x9]))
        # terminate column names list by 0x0
        file.write(bytes([0]))
        # write number of rows
        file.write((len(tabular_data_file.rows)).to_bytes(4, byteorder='little'))
        # write row indices, each terminated by tab (0x9)
        for row_name in tabular_data_file.rows:
            file.write(row_name.encode('utf-8'))
            file.write(bytes([0x9]))

def create_file(parsed):
    if (parsed.input.endswith('xlsx')):
        tabular_data_file = read_excel_file(parsed.input);
    else:
        print('unknown input file format')
        return

    if parsed.output:
        output_filename = parsed.output
    else:
        filename = os.path.split(parsed.input)[1]
        basename = os.path.splitext(filename)[0]
        output_filename = basename+'.2da'

    write_2da_file(output_filename, tabular_data_file)


def error(output_filename, tabular_data_file):
    print("You need to specify one of -x, -c")

def execute_action(parsed):
    function = parsed.action+"_"+parsed.format
    switcher= {
        "extract_csv" : extract_csv,
        "extract_excel" : extract_excel,
        "create_csv" : create_file,
        "create_excel" : create_file
    }
    func = switcher.get(function, error)
    func(parsed)


def parse_command_line():
    parser = argparse.ArgumentParser(description='Process 2DA files.')
    parser.add_argument('input', help='input file')
    parser.add_argument('output', nargs='?', help='output file (default: derive filename from input file and format)')
    parser.add_argument('-x', action='store_const', dest='action', const='extract', help='extract 2da file')
    parser.add_argument('-c', action='store_const', dest='action', const='create', help='create 2da file (Not Yet Implemented)')
    parser.add_argument('-f', choices=['csv', 'excel'], dest='format', default='csv', help='input format: csv or excel. (default: csv)')
    parser.add_argument('--csvsep', nargs='?', const=',', default=',', help="set csv separator (default: comma ',')")

    parsed = parser.parse_args()
    execute_action(parsed)


def main():
    parse_command_line()

main()

